import io
from datetime import datetime


def export_excel(products):
    """Export products to Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory Report"

        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = "INVENTORY MANAGEMENT REPORT"
        ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="1a2a4a")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:G2")
        ws["A2"] = f"Generated on: {datetime.now().strftime('%d %B %Y, %I:%M %p')}"
        ws["A2"].alignment = Alignment(horizontal="center")
        ws["A2"].font = Font(italic=True, size=10)

        # Headers
        headers = ["ID", "Product Name", "Category", "Price (Rs)", "Quantity", "Status", "Description"]
        header_fill = PatternFill("solid", fgColor="0d7377")
        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin

        # Data rows
        status_colors = {
            "In Stock":     "d4edda",
            "Low Stock":    "fff3cd",
            "Out of Stock": "f8d7da"
        }

        for row_idx, p in enumerate(products, 5):
            data = [p[0], p[1], p[2], f"{p[3]:,.2f}", p[4], p[5], p[6]]
            status = p[5]
            row_fill = PatternFill("solid", fgColor=status_colors.get(status, "FFFFFF"))

            for col_idx, val in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = row_fill
                cell.border = thin
                cell.alignment = Alignment(horizontal="center" if col_idx in [1,4,5] else "left")

        # Column widths
        widths = [6, 25, 15, 14, 10, 15, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Summary
        last_row = len(products) + 6
        ws.cell(row=last_row, column=1, value="SUMMARY").font = Font(bold=True)
        total_val = sum(p[3] * p[4] for p in products)
        ws.cell(row=last_row, column=2, value=f"Total Products: {len(products)}")
        ws.cell(row=last_row, column=4, value=f"Total Value: Rs {total_val:,.2f}").font = Font(bold=True)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    except ImportError:
        # Fallback: plain CSV in memory
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["ID", "Name", "Category", "Price", "Quantity", "Status", "Description"])
        for p in products:
            writer.writerow(p)
        output.seek(0)
        return io.BytesIO(output.getvalue().encode())


def export_pdf(products, stats):
    """Export inventory report as PDF."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        NAVY  = colors.HexColor("#1a2a4a")
        TEAL  = colors.HexColor("#0d7377")
        GREEN = colors.HexColor("#d4edda")
        AMBER = colors.HexColor("#fff3cd")
        RED   = colors.HexColor("#f8d7da")

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4),
                                rightMargin=15*mm, leftMargin=15*mm,
                                topMargin=15*mm, bottomMargin=15*mm)

        title_style = ParagraphStyle("title", fontName="Helvetica-Bold",
                                     fontSize=18, textColor=NAVY, alignment=TA_CENTER, spaceAfter=4)
        sub_style   = ParagraphStyle("sub",   fontName="Helvetica",
                                     fontSize=10, textColor=colors.gray, alignment=TA_CENTER, spaceAfter=12)

        story = [
            Paragraph("INVENTORY MANAGEMENT REPORT", title_style),
            Paragraph(f"Generated: {datetime.now().strftime('%d %B %Y, %I:%M %p')} &nbsp;|&nbsp; "
                      f"Total Products: {stats['total']} &nbsp;|&nbsp; "
                      f"Inventory Value: Rs {stats['total_value']:,.2f}", sub_style),
        ]

        # Stats row
        stat_data = [[
            f"In Stock: {stats['in_stock']}",
            f"Low Stock: {stats['low_stock']}",
            f"Out of Stock: {stats['out_of_stock']}",
            f"Total Value: Rs {stats['total_value']:,.2f}"
        ]]
        stat_table = Table(stat_data, colWidths=[65*mm]*4)
        stat_table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (0,0), GREEN),
            ("BACKGROUND", (1,0), (1,0), AMBER),
            ("BACKGROUND", (2,0), (2,0), RED),
            ("BACKGROUND", (3,0), (3,0), colors.HexColor("#cce5ff")),
            ("ALIGN",      (0,0), (-1,-1), "CENTER"),
            ("FONTNAME",   (0,0), (-1,-1), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0), (-1,-1), 11),
            ("BOX",        (0,0), (-1,-1), 1, TEAL),
            ("INNERGRID",  (0,0), (-1,-1), 0.5, colors.white),
            ("TOPPADDING", (0,0), (-1,-1), 8),
            ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ]))
        story.append(stat_table)
        story.append(Spacer(1, 10*mm))

        # Product table
        headers = [["#", "Product Name", "Category", "Price (Rs)", "Qty", "Status"]]
        rows = []
        for p in products:
            rows.append([str(p[0]), p[1], p[2], f"{p[3]:,.2f}", str(p[4]), p[5]])

        table_data = headers + rows
        col_widths = [12*mm, 65*mm, 35*mm, 30*mm, 15*mm, 35*mm]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)

        style = TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), NAVY),
            ("TEXTCOLOR",     (0,0), (-1,0), colors.white),
            ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",      (0,0), (-1,0), 10),
            ("ALIGN",         (0,0), (-1,-1), "CENTER"),
            ("ALIGN",         (1,1), (2,-1), "LEFT"),
            ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
            ("FONTSIZE",      (0,1), (-1,-1), 9),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, colors.HexColor("#f8f9fa")]),
            ("BOX",           (0,0), (-1,-1), 1, TEAL),
            ("INNERGRID",     (0,0), (-1,-1), 0.5, colors.HexColor("#dee2e6")),
            ("TOPPADDING",    (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ])

        # Color rows by status
        for i, p in enumerate(products, 1):
            if p[5] == "Out of Stock":
                style.add("BACKGROUND", (0,i), (-1,i), RED)
            elif p[5] == "Low Stock":
                style.add("BACKGROUND", (0,i), (-1,i), AMBER)

        table.setStyle(style)
        story.append(table)

        doc.build(story)
        output.seek(0)
        return output

    except ImportError:
        output = io.BytesIO(b"PDF export requires: pip install reportlab")
        output.seek(0)
        return output
